from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_calendar_months(output_path="calendar_months.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendar Months"

    months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    # Header
    ws['A1'] = "Month"
    ws['B1'] = "Month Number"
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='4472C4')
    for cell in ['A1', 'B1']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = Alignment(horizontal='center')

    # Data
    for i, month in enumerate(months, start=2):
        ws[f'A{i}'] = month
        ws[f'B{i}'] = i - 1
        ws[f'A{i}'].font = Font(name='Arial')
        ws[f'B{i}'].font = Font(name='Arial')
        ws[f'B{i}'].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")

if __name__ == "__main__":
    generate_calendar_months()
